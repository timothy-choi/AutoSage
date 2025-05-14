import os
from typing import List, Optional
from PyPDF2 import PdfReader, PdfWriter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
import textract # type: ignore
import shutil
import datetime
import openpyxl
from openpyxl import Workbook, load_workbook

def create_excel(data: List[List[str]], output_path: str) -> str:
    wb = Workbook()
    ws = wb.active
    for row in data:
        ws.append(row)
    wb.save(output_path)
    return output_path

def read_excel(file_path: str) -> List[List[str]]:
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
    wb = load_workbook(file_path)
    ws = wb.active
    return [[str(cell.value) if cell.value is not None else "" for cell in row] for row in ws.iter_rows()]

def append_row(file_path: str, row_data: List[str]) -> str:
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
    wb = load_workbook(file_path)
    ws = wb.active
    ws.append(row_data)
    wb.save(file_path)
    return file_path

def get_cell_value(file_path: str, cell: str) -> str:
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
    wb = load_workbook(file_path)
    ws = wb.active
    value = ws[cell].value
    return str(value) if value is not None else ""

def set_cell_value(file_path: str, cell: str, value: str) -> str:
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
    wb = load_workbook(file_path)
    ws = wb.active
    ws[cell] = value
    wb.save(file_path)
    return file_path

def delete_row(file_path: str, row: int) -> str:
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
    wb = load_workbook(file_path)
    ws = wb.active
    ws.delete_rows(row)
    wb.save(file_path)
    return file_path

def copy_excel(source_path: str, destination_path: str) -> str:
    if not os.path.exists(source_path):
        raise FileNotFoundError(f"File not found: {source_path}")
    shutil.copy(source_path, destination_path)
    return destination_path

def rename_excel(original_path: str, new_path: str) -> str:
    if not os.path.exists(original_path):
        raise FileNotFoundError(f"File not found: {original_path}")
    os.rename(original_path, new_path)
    return new_path

def delete_excel(file_path: str) -> None:
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
    os.remove(file_path)

def extract_text_from_excel(file_path: str) -> str:
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
    wb = load_workbook(file_path)
    ws = wb.active
    text = ""
    for row in ws.iter_rows(values_only=True):
        text += "\t".join([str(cell) if cell is not None else "" for cell in row]) + "\n"
    return text